import tempfile
from collections import OrderedDict

import pytest
from hypothesis import given, strategies as st, settings

import numpy as np
import pandas as pd

from ._testtools import strategy_2d_prob_distribution

from tmtoolkit.topicmod import model_io


def test_save_load_ldamodel_pickle():
    try:
        import lda
    except ImportError:
        pytest.skip('lda not installed')

    pfile = 'tests/data/test_pickle_unpickle_ldamodel.pickle'

    dtm = np.array([[0, 1], [2, 3], [4, 5], [6, 0]])
    doc_labels = ['doc_' + str(i) for i in range(dtm.shape[0])]
    vocab = ['word_' + str(i) for i in range(dtm.shape[1])]

    model = lda.LDA(2, n_iter=1)
    model.fit(dtm)

    model_io.save_ldamodel_to_pickle(pfile, model, vocab, doc_labels)

    unpickled = model_io.load_ldamodel_from_pickle(pfile)

    assert np.array_equal(model.doc_topic_, unpickled['model'].doc_topic_)
    assert np.array_equal(model.topic_word_, unpickled['model'].topic_word_)
    assert vocab == unpickled['vocab']
    assert doc_labels == unpickled['doc_labels']


@given(
    topic_word=strategy_2d_prob_distribution(),
    top_n=st.integers(min_value=0, max_value=20)
)
@settings(deadline=1000)
def test_ldamodel_top_topic_words(topic_word, top_n):
    topic_word = np.array(topic_word)

    vocab = np.array(['t%d' % i for i in range(topic_word.shape[1])])

    if top_n < 1 or top_n > topic_word.shape[1]:
        with pytest.raises(ValueError):
            model_io.ldamodel_top_topic_words(topic_word, vocab, top_n)
    else:
        top_topic_words = model_io.ldamodel_top_topic_words(topic_word, vocab, top_n)
        colnames = np.array([model_io.DEFAULT_RANK_NAME_FMT.format(i1=i + 1) for i in range(top_n)])
        rownames = np.array([model_io.DEFAULT_TOPIC_NAME_FMT.format(i1=i + 1) for i in range(topic_word.shape[0])])

        assert top_topic_words.shape == (topic_word.shape[0], top_n)
        assert np.array_equal(top_topic_words.index.values, rownames)
        assert np.array_equal(top_topic_words.columns.values, colnames)


@given(
    topic_word=strategy_2d_prob_distribution(),
    top_n=st.integers(min_value=0, max_value=20)
)
@settings(deadline=1000)
def test_ldamodel_top_word_topics(topic_word, top_n):
    topic_word = np.array(topic_word)

    vocab = np.array(['t%d' % i for i in range(topic_word.shape[1])])

    if top_n < 1 or top_n > topic_word.shape[0]:
        with pytest.raises(ValueError):
            model_io.ldamodel_top_word_topics(topic_word, vocab, top_n)
    else:
        top_word_topics = model_io.ldamodel_top_word_topics(topic_word, vocab, top_n)
        colnames = np.array([model_io.DEFAULT_RANK_NAME_FMT.format(i1=i + 1) for i in range(top_n)])

        assert top_word_topics.shape == (topic_word.shape[1], top_n) == (len(vocab), top_n)
        assert np.array_equal(top_word_topics.index.values, vocab)
        assert np.array_equal(top_word_topics.columns.values, colnames)


@given(
    doc_topic=strategy_2d_prob_distribution(),
    top_n=st.integers(min_value=0, max_value=20)
)
@settings(deadline=1000)
def test_ldamodel_top_doc_topics(doc_topic, top_n):
    doc_topic = np.array(doc_topic)

    doc_labels = np.array(['doc%d' % i for i in range(doc_topic.shape[0])])

    if top_n < 1 or top_n > doc_topic.shape[1]:
        with pytest.raises(ValueError):
            model_io.ldamodel_top_topic_words(doc_topic, doc_labels, top_n)
    else:
        top_doc_topics = model_io.ldamodel_top_doc_topics(doc_topic, doc_labels, top_n)
        colnames = np.array([model_io.DEFAULT_RANK_NAME_FMT.format(i1=i + 1) for i in range(top_n)])

        assert top_doc_topics.shape == (doc_topic.shape[0], top_n)
        assert np.array_equal(top_doc_topics.index.values, doc_labels)
        assert np.array_equal(top_doc_topics.columns.values, colnames)


@given(
    doc_topic=strategy_2d_prob_distribution(),
    top_n=st.integers(min_value=0, max_value=20)
)
@settings(deadline=1000)
def test_ldamodel_top_topic_docs(doc_topic, top_n):
    doc_topic = np.array(doc_topic)

    doc_labels = np.array(['doc%d' % i for i in range(doc_topic.shape[0])])

    if top_n < 1 or top_n > doc_topic.shape[0]:
        with pytest.raises(ValueError):
            model_io.ldamodel_top_topic_docs(doc_topic, doc_labels, top_n)
    else:
        top_topic_docs = model_io.ldamodel_top_topic_docs(doc_topic, doc_labels, top_n)
        colnames = np.array([model_io.DEFAULT_RANK_NAME_FMT.format(i1=i + 1) for i in range(top_n)])
        rownames = np.array([model_io.DEFAULT_TOPIC_NAME_FMT.format(i1=i + 1) for i in range(doc_topic.shape[1])])

        assert top_topic_docs.shape == (doc_topic.shape[1], top_n)
        assert np.array_equal(top_topic_docs.index.values, rownames)
        assert np.array_equal(top_topic_docs.columns.values, colnames)


@given(topic_word=strategy_2d_prob_distribution())
@settings(deadline=1000)
def test_ldamodel_full_topic_words(topic_word):
    topic_word = np.array(topic_word)

    vocab = np.array(['t%d' % i for i in range(topic_word.shape[1])])

    df = model_io.ldamodel_full_topic_words(topic_word, vocab)
    assert isinstance(df, pd.DataFrame)

    rownames = np.array([model_io.DEFAULT_TOPIC_NAME_FMT.format(i1=i + 1) for i in range(topic_word.shape[0])])
    assert df.columns.tolist() == ['_topic'] + list(vocab)

    assert np.array_equal(df.iloc[:, 0].to_numpy(), rownames)


@given(doc_topic=strategy_2d_prob_distribution())
@settings(deadline=1000)
def test_ldamodel_full_doc_topics(doc_topic):
    doc_topic = np.array(doc_topic)

    doc_labels = np.array(['doc%d' % i for i in range(doc_topic.shape[0])])

    df = model_io.ldamodel_full_doc_topics(doc_topic, doc_labels)
    assert isinstance(df, pd.DataFrame)

    colnames = np.array([model_io.DEFAULT_TOPIC_NAME_FMT.format(i1=i + 1) for i in range(doc_topic.shape[1])])
    assert df.columns.tolist() == ['_doc'] + list(colnames)

    assert np.array_equal(df.iloc[:, 0].to_numpy(), doc_labels)


@settings(deadline=10000)
@given(n_docs=st.integers(min_value=0, max_value=10),
       n_topics=st.integers(min_value=0, max_value=10),
       size_vocab=st.integers(min_value=0, max_value=50),
       top_n_topics=st.integers(min_value=0, max_value=10),
       top_n_words=st.integers(min_value=0, max_value=50),
       create_dtm=st.booleans())
def test_save_ldamodel_summary_to_excel(n_docs, n_topics, size_vocab, top_n_topics, top_n_words, create_dtm):
    try:
        import openpyxl
    except ImportError:
        pytest.skip('openpyxl not installed')

    topic_word = np.random.uniform(size=n_topics * size_vocab).reshape((n_topics, size_vocab))
    doc_topic = np.random.uniform(size=n_docs * n_topics).reshape((n_docs, n_topics))
    doc_labels = np.array(['doc%d' % i for i in range(doc_topic.shape[0])])
    vocab = np.array(['t%d' % i for i in range(topic_word.shape[1])])
    _, excelfile = tempfile.mkstemp(suffix='.xlsx')

    if create_dtm:
        dtm = np.random.randint(0, 10, size=n_docs*size_vocab).reshape(n_docs, size_vocab)
    else:
        dtm = None

    if top_n_words < 1 or top_n_words > topic_word.shape[1] or top_n_topics < 1 or top_n_topics > topic_word.shape[0]\
            or n_docs < 1:
        with pytest.raises(ValueError):
            model_io.save_ldamodel_summary_to_excel(excelfile, topic_word, doc_topic, doc_labels, vocab,
                                                    top_n_topics=top_n_topics, top_n_words=top_n_words)
    else:
        excelsheets = model_io.save_ldamodel_summary_to_excel(excelfile, topic_word, doc_topic, doc_labels, vocab,
                                                              top_n_topics=top_n_topics, top_n_words=top_n_words,
                                                              dtm=dtm)
        assert isinstance(excelsheets, OrderedDict)

        sheetnames = ['top_doc_topics_vals', 'top_doc_topics_labels', 'top_doc_topics_labelled_vals',
                      'top_topic_word_vals', 'top_topic_word_labels', 'top_topic_words_labelled_vals']

        if dtm is not None:
            sheetnames.append('marginal_topic_distrib')

        assert list(excelsheets.keys()) == sheetnames

        for sheetn in sheetnames:
            assert isinstance(excelsheets[sheetn], pd.DataFrame)

